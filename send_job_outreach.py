#!/usr/bin/env python3
"""
Send personalized job outreach emails from a CSV or Excel contacts file.

Contact file columns:
  email, company, name

The "name" column is optional. The script previews emails by default; add
--send only when you are ready to actually send them.
"""

from __future__ import annotations

import argparse
import csv
import mimetypes
import os
import re
import smtplib
import ssl
import sys
import time
from dataclasses import dataclass
from email.message import EmailMessage
from email.utils import formataddr, make_msgid
from pathlib import Path


DEFAULT_RESUME_PATH = Path(r"D:\Resume\FE\Achyutananda_Nayak_Resume.pdf")
DEFAULT_SUBJECT = "Interest in Full Stack Opportunities at {company}"
DEFAULT_SENDER_NAME = "Achyutananda Nayak"
DEFAULT_SMTP_HOST = "smtp.gmail.com"
DEFAULT_SMTP_PORT = 587

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

BODY_TEMPLATE = """Hi {name},

I hope you're doing well.

I came across your profile while exploring opportunities at {company} and wanted to reach out regarding any Full Stack Software Engineer openings that may be available.

I have 4 years of experience in software development, primarily working with React, Next.js, JavaScript, TypeScript, Node.js, Redux, AWS, and microservices. Throughout my career, I've built scalable web applications, REST APIs, enterprise dashboards, and performance-focused user experiences.

I'm currently exploring opportunities where I can contribute to impactful products while continuing to grow as an engineer. Based on my background, I believe I could be a strong fit for Full Stack roles at {company}.

If there are any relevant openings, I would be grateful if you could consider my profile or guide me through the appropriate application process. I've attached my resume for your review.

Thank you for your time and consideration.

Best regards,
Achyutananda Nayak
"""


@dataclass(frozen=True)
class Contact:
    email: str
    company: str
    name: str
    row_number: int


def clean_value(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def first_value(row: dict[str, str], *names: str) -> str:
    for name in names:
        value = row.get(name)
        if value:
            return value.strip()
    return ""


def normalize_row(row: dict[object, object]) -> dict[str, str]:
    return {clean_value(key).lower(): clean_value(value) for key, value in row.items()}


def load_contacts(path: Path) -> list[Contact]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        rows = load_xlsx_rows(path)
    elif suffix == ".csv":
        rows = load_csv_rows(path)
    else:
        raise ValueError(f"Unsupported contacts file type: {path.suffix}. Use .csv or .xlsx.")

    contacts: list[Contact] = []
    problems: list[str] = []

    for row_number, raw_row in rows:
        row = normalize_row(raw_row)
        email = first_value(row, "email", "email_id", "mail", "mail_id")
        company = first_value(row, "company", "company_name", "organisation", "organization")
        name = first_value(row, "name", "recipient_name", "first_name") or "there"

        if not email or not company:
            problems.append(f"row {row_number}: missing email or company")
            continue

        if not EMAIL_RE.match(email):
            problems.append(f"row {row_number}: invalid email address {email!r}")
            continue

        contacts.append(Contact(email=email, company=company, name=name, row_number=row_number))

    if problems:
        print("Skipped rows:", file=sys.stderr)
        for problem in problems:
            print(f"  - {problem}", file=sys.stderr)

    return contacts


def load_csv_rows(path: Path) -> list[tuple[int, dict[object, object]]]:
    rows: list[tuple[int, dict[object, object]]] = []

    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        for row_number, raw_row in enumerate(reader, start=2):
            rows.append((row_number, raw_row))

    return rows


def load_xlsx_rows(path: Path) -> list[tuple[int, dict[object, object]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Excel contacts require openpyxl. Use .csv or install openpyxl.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)

    try:
        headers = next(rows_iter)
    except StopIteration:
        return []

    header_names = [clean_value(header) for header in headers]
    rows: list[tuple[int, dict[object, object]]] = []

    for row_number, values in enumerate(rows_iter, start=2):
        if not any(clean_value(value) for value in values):
            continue
        rows.append((row_number, dict(zip(header_names, values))))

    workbook.close()
    return rows


def add_attachment(message: EmailMessage, attachment_path: Path) -> None:
    content_type, _ = mimetypes.guess_type(attachment_path.name)
    maintype, subtype = (content_type or "application/octet-stream").split("/", 1)

    with attachment_path.open("rb") as file:
        message.add_attachment(
            file.read(),
            maintype=maintype,
            subtype=subtype,
            filename=attachment_path.name,
        )


def build_message(
    contact: Contact,
    *,
    from_email: str,
    sender_name: str,
    subject_template: str,
    resume_path: Path,
    test_to: str | None,
) -> EmailMessage:
    subject = subject_template.format(company=contact.company, name=contact.name)
    body = BODY_TEMPLATE.format(company=contact.company, name=contact.name)
    to_email = test_to or contact.email

    message = EmailMessage()
    message["From"] = formataddr((sender_name, from_email))
    message["To"] = to_email
    message["Subject"] = subject
    message["Message-ID"] = make_msgid()

    if test_to:
        message["X-Original-To"] = contact.email

    message.set_content(body)
    add_attachment(message, resume_path)
    return message


def normalize_smtp_password(password: str, smtp_host: str) -> str:
    if "gmail" in smtp_host.lower():
        return "".join(password.split())
    return password


def preview_message(contact: Contact, message: EmailMessage, resume_path: Path) -> None:
    print("=" * 72)
    print(f"Contact row: {contact.row_number}")
    print(f"To: {message['To']}")
    if message.get("X-Original-To"):
        print(f"Original To: {message['X-Original-To']}")
    print(f"Subject: {message['Subject']}")
    print(f"Attachment: {resume_path}")
    print("-" * 72)
    print(get_plain_text_body(message).strip())
    print()


def get_plain_text_body(message: EmailMessage) -> str:
    if not message.is_multipart():
        return message.get_content()

    for part in message.walk():
        if part.get_content_type() == "text/plain" and part.get_content_disposition() != "attachment":
            return part.get_content()

    return ""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Send personalized job outreach emails.")
    parser.add_argument(
        "--contacts",
        type=Path,
        default=Path("contacts.csv"),
        help="CSV or XLSX file with email, company, and optional name columns.",
    )
    parser.add_argument(
        "--resume",
        type=Path,
        default=DEFAULT_RESUME_PATH,
        help="Resume PDF path to attach.",
    )
    parser.add_argument("--send", action="store_true", help="Actually send the emails.")
    parser.add_argument("--limit", type=int, default=None, help="Only process the first N valid contacts.")
    parser.add_argument("--delay", type=float, default=2.0, help="Seconds to wait between sent emails.")
    parser.add_argument("--test-to", default=None, help="Send every generated email to this address for testing.")
    parser.add_argument("--subject", default=DEFAULT_SUBJECT, help="Subject template. You can use {company}.")
    parser.add_argument("--smtp-host", default=os.getenv("SMTP_HOST", DEFAULT_SMTP_HOST))
    parser.add_argument("--smtp-port", type=int, default=int(os.getenv("SMTP_PORT", DEFAULT_SMTP_PORT)))
    parser.add_argument("--from-email", default=os.getenv("SMTP_EMAIL"), help="Sender email address.")
    parser.add_argument(
        "--password",
        default=os.getenv("SMTP_APP_PASSWORD") or os.getenv("SMTP_PASSWORD"),
        help="SMTP password or Gmail app password. Prefer setting SMTP_APP_PASSWORD instead of typing it here.",
    )
    parser.add_argument("--sender-name", default=os.getenv("SENDER_NAME", DEFAULT_SENDER_NAME))
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    contacts_path = args.contacts.expanduser()
    resume_path = args.resume.expanduser()

    if not contacts_path.exists():
        print(f"Contacts file not found: {contacts_path}", file=sys.stderr)
        return 1

    if not resume_path.exists():
        print(f"Resume attachment not found: {resume_path}", file=sys.stderr)
        return 1

    try:
        contacts = load_contacts(contacts_path)
    except Exception as exc:
        print(f"Could not load contacts: {exc}", file=sys.stderr)
        return 1
    if args.limit is not None:
        contacts = contacts[: args.limit]

    if not contacts:
        print("No valid contacts found.", file=sys.stderr)
        return 1

    messages = [
        build_message(
            contact,
            from_email=args.from_email or "preview@example.com",
            sender_name=args.sender_name,
            subject_template=args.subject,
            resume_path=resume_path,
            test_to=args.test_to,
        )
        for contact in contacts
    ]

    if not args.send:
        print("Preview mode only. Add --send when you are ready to send.\n")
        for contact, message in zip(contacts, messages):
            preview_message(contact, message, resume_path)
        print(f"Prepared {len(messages)} email(s). Nothing was sent.")
        return 0

    if not args.from_email:
        print("Missing sender email. Set SMTP_EMAIL or pass --from-email.", file=sys.stderr)
        return 1

    if not args.password:
        print("Missing SMTP password. Set SMTP_APP_PASSWORD or pass --password.", file=sys.stderr)
        return 1

    context = ssl.create_default_context()
    sent = 0
    failed = 0

    with smtplib.SMTP(args.smtp_host, args.smtp_port) as smtp:
        smtp.ehlo()
        smtp.starttls(context=context)
        smtp.ehlo()
        smtp_password = normalize_smtp_password(args.password, args.smtp_host)

        try:
            smtp.login(args.from_email, smtp_password)
        except smtplib.SMTPAuthenticationError:
            print("Gmail rejected the login.", file=sys.stderr)
            print("Use a 16-character Gmail app password, not your normal Gmail password.", file=sys.stderr)
            print("Create one here: https://myaccount.google.com/apppasswords", file=sys.stderr)
            return 1

        for index, (contact, message) in enumerate(zip(contacts, messages), start=1):
            try:
                smtp.send_message(message)
                sent += 1
                print(f"[{index}/{len(messages)}] Sent to {message['To']} for {contact.company}")
            except Exception as exc:
                failed += 1
                print(f"[{index}/{len(messages)}] Failed for {contact.email}: {exc}", file=sys.stderr)

            if index < len(messages) and args.delay > 0:
                time.sleep(args.delay)

    print(f"Done. Sent: {sent}. Failed: {failed}.")
    return 0 if failed == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
