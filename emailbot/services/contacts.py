"""Loading and validation of contacts from CSV or XLSX files."""

from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

from emailbot.models import Contact
from emailbot.services.jobids import extract_job_ids

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
YES_VALUES = {"1", "true", "yes", "y"}
FRONTEND_OUTREACH_COLUMNS = (
    "frontend_outreach",
    "frontend outreach",
    "frontend_outreach_yes",
    "frontend outreach yes",
    "frontend",
)


def clean_value(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def first_value(row: dict[str, str], *names: str) -> str:
    for name in names:
        value = row.get(name)
        if value:
            return value.strip()
    return ""


def is_yes(value: str) -> bool:
    return value.strip().lower() in YES_VALUES


def normalize_row(row: dict[object, object]) -> dict[str, str]:
    return {clean_value(key).lower(): clean_value(value) for key, value in row.items()}


def load_contacts(path: Path) -> list[Contact]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        rows = load_xlsx_rows(path)
    elif suffix == ".csv":
        rows = load_csv_rows(path)
    else:
        raise ValueError(f"Unsupported contacts file type: {path.suffix}. Use .csv or .xlsx.")

    contacts: list[Contact] = []
    problems: list[str] = []

    for row_number, raw_row in rows:
        row = normalize_row(raw_row)
        email = first_value(row, "email", "email_id", "mail", "mail_id")
        company = first_value(row, "company", "company_name", "organisation", "organization")
        name = first_value(row, "name", "recipient_name", "first_name") or "there"
        job_ids = extract_job_ids(row)
        frontend_outreach = is_yes(first_value(row, *FRONTEND_OUTREACH_COLUMNS))

        if not email or not company:
            problems.append(f"row {row_number}: missing email or company")
            continue

        if not EMAIL_RE.match(email):
            problems.append(f"row {row_number}: invalid email address {email!r}")
            continue

        contacts.append(
            Contact(
                email=email,
                company=company,
                name=name,
                job_ids=job_ids,
                frontend_outreach=frontend_outreach,
                row_number=row_number,
            )
        )

    if problems:
        print("Skipped rows:", file=sys.stderr)
        for problem in problems:
            print(f"  - {problem}", file=sys.stderr)

    return contacts


def load_csv_rows(path: Path) -> list[tuple[int, dict[object, object]]]:
    rows: list[tuple[int, dict[object, object]]] = []

    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        for row_number, raw_row in enumerate(reader, start=2):
            rows.append((row_number, raw_row))

    return rows


def load_xlsx_rows(path: Path) -> list[tuple[int, dict[object, object]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Excel contacts require openpyxl. Use .csv or install openpyxl.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)

    try:
        headers = next(rows_iter)
    except StopIteration:
        return []

    header_names = [clean_value(header) for header in headers]
    rows: list[tuple[int, dict[object, object]]] = []

    for row_number, values in enumerate(rows_iter, start=2):
        if not any(clean_value(value) for value in values):
            continue
        rows.append((row_number, dict(zip(header_names, values))))

    workbook.close()
    return rows
